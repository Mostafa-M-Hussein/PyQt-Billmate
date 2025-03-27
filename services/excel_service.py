from typing import List

from PyQt5.QtWidgets import QMessageBox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo, TableColumn, AutoFilter


class ExcelService:
    def create_excel_sheet(
        self,
        headers: List[str],
        data_rows: List[List[str]],  # Corrected type to List[List[str]]
        path: str = "test.xlsx",
    ):
        wb = Workbook()
        ws = wb.active
        try:
            # Headers for the table
            ws.sheet_view.rightToLeft = True
            ws.append([col.strip() for col in headers])

            # Append the rows of data
            for row in data_rows:
                print(row)
                ws.append([str(col) for col in row])

            # Adjust column widths
            for col_num, col_title in enumerate(headers, 1):
                column_letter = get_column_letter(col_num)
                max_length = max(len(col_title), 15)
                for row in data_rows:
                    if len(row) >= col_num:  # Ensure the row has enough columns
                        max_length = max(max_length, len(str(row[col_num - 1])))
                ws.column_dimensions[column_letter].width = (
                    max_length + 2
                )  # Adding padding

            # Apply right alignment to all cells
            alignment = Alignment(horizontal="right")
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = alignment

            # Define table range
            table_start_row = 1
            table_start_col = 1
            table_end_row = table_start_row + len(
                data_rows
            )  # Corrected end row calculation
            table_end_col = len(headers)

            # Apply alternating row colors
            fill1 = PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
            )
            fill2 = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )
            for idx, row in enumerate(
                ws.iter_rows(min_row=table_start_row + 1), start=1
            ):
                fill = fill1 if idx % 2 == 0 else fill2
                for cell in row:
                    cell.fill = fill

            # Create table reference and columns
            ref = f"{get_column_letter(table_start_col)}{table_start_row}:{get_column_letter(table_end_col)}{table_end_row}"
            columns = [
                TableColumn(
                    id=col_id,
                    name=col_header,
                    totalsRowFunction="sum" if col_id == len(headers) else None,
                )
                for col_id, col_header in enumerate(headers, start=1)
            ]

            # Define table style
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=False,
                showColumnStripes=True,
            )

            tab = Table(
                displayName="Table1",
                ref=ref,
                tableStyleInfo=style,
                totalsRowCount=1 if len(data_rows) > 0 else 0,
                totalsRowShown=True if len(data_rows) > 0 else False,
                tableColumns=columns,
                autoFilter=AutoFilter(ref=ref),
            )
            ws.add_table(tab)
            wb.save(path)
        except Exception as e:
            QMessageBox.warning(None, "Export Failed", str(e))
            return False
        finally:
            wb.close()
        return True
